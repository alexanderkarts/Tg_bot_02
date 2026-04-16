import json
import logging
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, CommandHandler, ContextTypes, MessageHandler, filters
)
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from config import TELEGRAM_TOKEN, SHEET_URL, GOOGLE_CREDENTIALS_JSON, GOOGLE_SHEET_URL
from utils import full_stock, PHOTO_COLUMNS, NOT_SENT_COLUMNS

logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

# ===== Настройка кнопок =====
reply_keyboard = [
    ["Статистика"],
    ["Полный сток", "Авто без фото"],
    ["Авто без места хранения"],
    ["Переданные авто", "Не переданные авто"],
    ["Поиск ключа"]
]
markup = ReplyKeyboardMarkup(reply_keyboard, resize_keyboard=True)

# ===== Google Sheets (lazy) =====
_gs_sheet = None

def get_sheet():
    global _gs_sheet
    if _gs_sheet is None:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds_value = GOOGLE_CREDENTIALS_JSON
        if creds_value.strip().startswith("{"):
            creds_dict = json.loads(creds_value)
            creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        else:
            creds = ServiceAccountCredentials.from_json_keyfile_name(creds_value, scope)
        client = gspread.authorize(creds)
        _gs_sheet = client.open_by_url(GOOGLE_SHEET_URL).sheet1
    return _gs_sheet

# ===== Кеш ключей =====
KEYS_FILE = "keys.json"

def load_keys() -> dict:
    try:
        with open(KEYS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return {}

def save_keys(keys_data: dict):
    with open(KEYS_FILE, "w", encoding="utf-8") as f:
        json.dump(keys_data, f, ensure_ascii=False, indent=2)

# ===== Глобальные данные =====
WAITING_KEY = {}

SENT_COLUMNS = [
    "Номер ключа", "VIN", "Кол-во фото для сайта", "Модель", "Марка",
    "Пробег", "Год выпуска", "Цветкузова", "Рег. номер",
    "Дней с даты поступления", "ДЦ приёма", "Тип сделки"
]

# ===== Извлечение номеров ключей из текста =====
def extract_key_numbers(text: str) -> list:
    if not text:
        return []
    t = text.lower()
    found = []

    # A: "ключ N" — число ПОСЛЕ слова "ключ"
    # Джетта 4553- ключ 285 / Мондео 1288- ключ 3
    for m in re.finditer(r'ключ\s*[:\-]?\s*(\d{1,4})(?!\d)', t):
        found.append(int(m.group(1)))

    # B: "N ключ" — число ДО слова "ключ" (не ключа/ключей — это количество)
    # 370 ключ (2 шт), 364 ключ
    for m in re.finditer(r'(?<!\d)(\d{1,4})\s*[-.]?\s*ключ(?!а|ей|ами|\w)', t):
        num = int(m.group(1))
        if num not in found:
            found.append(num)

    # C: количество кл/ключа/ключей потом номер в скобках: "2кл (46)", "2 ключа (46)"
    for m in re.finditer(r'\d+\s*(?:ключей|ключа|кл(?!ю))\s*\((\d{1,4})\)', t):
        num = int(m.group(1))
        if num not in found:
            found.append(num)

    # D: "N (Mкл)" — номер ключа потом количество в скобках: "147 (2кл)"
    for m in re.finditer(r'(?<!\d)(\d{1,4})\s*\(\d+\s*кл\)', t):
        num = int(m.group(1))
        if num not in found:
            found.append(num)

    # E: "N. Выкуп" или "N выкуп" — номер в начале строки перед "выкуп"
    # 160. Выкуп 2кл
    for m in re.finditer(r'(?:^|\n)\s*(\d{1,4})\s*[.\-,]?\s*выкуп', t):
        num = int(m.group(1))
        if num not in found:
            found.append(num)

    return found

# ===== Загрузка данных =====
def load_data() -> pd.DataFrame:
    try:
        df = pd.read_csv(SHEET_URL)
        logger.info("Данные загружены")
        return df
    except Exception as e:
        logger.error(f"Ошибка загрузки данных: {e}")
        return pd.DataFrame()

# ===== Форматирование Excel =====
def format_excel(file_path: str):
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
    wb.save(file_path)

# ===== Запись DataFrame в Google Sheets =====
def write_df_to_sheet(sheet, df: pd.DataFrame):
    df_clean = df.fillna("")
    values = [df_clean.columns.tolist()] + df_clean.values.tolist()
    sheet.clear()
    sheet.update(range_name="A1", values=values)

# ===== Команды =====
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Привет! Выберите действие:", reply_markup=markup)

# ===== Главная обработка текста =====
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text or update.message.caption
    chat_type = update.message.chat.type

    logger.info(f"Чат: {chat_type} | текст: {text!r}")

    df = load_data()
    if df.empty:
        await update.message.reply_text("Ошибка: нет данных для отображения.")
        return

    # ===== Кнопки меню =====
    if text == "Полный сток":
        df_full = full_stock(df)
        file_path = "full_stock.xlsx"
        df_full.to_excel(file_path, index=False)
        format_excel(file_path)
        write_df_to_sheet(get_sheet(), df_full)
        await update.message.reply_text(f"Полный сток готов. Всего машин: {len(df_full)}")
        await update.message.reply_document(open(file_path, "rb"))
        return

    elif text == "Авто без фото":
        df_photo = df[df["Кол-во фото для сайта"] == 0]
        df_photo = df_photo.sort_values(by="Дней с даты поступления", ascending=False)
        df_photo = df_photo[PHOTO_COLUMNS]
        file_path = "auto_without_photo.xlsx"
        df_photo.to_excel(file_path, index=False)
        format_excel(file_path)
        write_df_to_sheet(get_sheet(), df_photo)
        await update.message.reply_text(f"Авто без фото готово. Всего машин: {len(df_photo)}")
        await update.message.reply_document(open(file_path, "rb"))
        return

    elif text == "Авто без места хранения":
        df_full = full_stock(df)
        df_no_storage = df_full[
            (df_full["Место хранения"].isna() | (df_full["Место хранения"] == "")) &
            (df_full["Кол-во фото для сайта"] != 0)
        ]
        df_no_storage = df_no_storage.sort_values(by="Дней с даты поступления", ascending=True)
        file_path = "Авто_без_места.xlsx"
        df_no_storage.to_excel(file_path, index=False)
        format_excel(file_path)
        write_df_to_sheet(get_sheet(), df_no_storage)
        await update.message.reply_document(open(file_path, "rb"), caption=f"Авто без места хранения: {len(df_no_storage)} машин")
        return

    elif text == "Статистика":
        df_full = full_stock(df)
        total_stock = len(df_full)
        no_photo = len(df_full[df_full["Кол-во фото для сайта"] == 0])
        no_storage = len(df_full[
            (df_full["Место хранения"].isna() | (df_full["Место хранения"] == "")) &
            (df_full["Кол-во фото для сайта"] != 0)
        ])
        parking_counts = df_full["Место хранения"].fillna("Без места").value_counts()
        message = f"📊 Статистика\n\nПолный сток: {total_stock}\nАвто без фото: {no_photo}\nАвто без места хранения: {no_storage}\n\nПарковки:\n"
        for parking, count in parking_counts.items():
            message += f"{parking}: {count}\n"
        await update.message.reply_text(message)
        return

    elif text == "Переданные авто":
        keys_data = load_keys()
        df_full = full_stock(df)
        # Переданными считаем только те авто, у которых совпадают И ключ, И VIN
        def is_truly_sent(row):
            key = str(int(row["Номер ключа"])) if pd.notna(row["Номер ключа"]) else ""
            entry = keys_data.get(key)
            if not entry or entry.get("status") != "sent":
                return False
            stored_vin = str(entry.get("VIN", "")).strip()
            current_vin = str(row.get("VIN", "")).strip()
            return stored_vin == current_vin and stored_vin != ""
        df_sent = df_full[df_full.apply(is_truly_sent, axis=1)]
        # Убираем авто у которых уже есть фото (фотографы обработали)
        df_sent = df_sent[df_sent["Кол-во фото для сайта"] == 0]
        # Сортировка по дням от максимального
        df_sent = df_sent.sort_values(by="Дней с даты поступления", ascending=False)
        # Оставляем только нужные колонки
        available = [c for c in SENT_COLUMNS if c in df_sent.columns]
        df_sent = df_sent[available]
        if df_sent.empty:
            await update.message.reply_text("Нет переданных авто.")
            return
        file_path = "Переданные_авто.xlsx"
        df_sent.to_excel(file_path, index=False)
        format_excel(file_path)
        write_df_to_sheet(get_sheet(), df_sent)
        await update.message.reply_document(open(file_path, "rb"), caption=f"Переданные авто ({len(df_sent)})")
        return

    elif text == "Не переданные авто":
        keys_data = load_keys()
        df_photo = df[df["Кол-во фото для сайта"] == 0].copy()
        # Не переданными считаем авто, у которых ключ не зарегистрирован
        # ИЛИ VIN в keys.json не совпадает с текущим (ключ переприсвоен новой машине)
        def is_not_sent(row):
            key = str(int(row["Номер ключа"])) if pd.notna(row["Номер ключа"]) else ""
            entry = keys_data.get(key)
            if not entry or entry.get("status") != "sent":
                return True
            stored_vin = str(entry.get("VIN", "")).strip()
            current_vin = str(row.get("VIN", "")).strip()
            return stored_vin != current_vin or stored_vin == ""
        df_not_sent = df_photo[df_photo.apply(is_not_sent, axis=1)]
        available = [c for c in NOT_SENT_COLUMNS if c in df_not_sent.columns]
        df_not_sent = df_not_sent[available]
        file_path = "Не_переданные_авто.xlsx"
        df_not_sent.to_excel(file_path, index=False)
        format_excel(file_path)
        write_df_to_sheet(get_sheet(), df_not_sent)
        await update.message.reply_document(open(file_path, "rb"), caption=f"Не переданные авто ({len(df_not_sent)})")
        return

    elif text == "Поиск ключа":
        WAITING_KEY[update.message.chat_id] = True
        await update.message.reply_text("Введите номер ключа")
        return

    elif WAITING_KEY.get(update.message.chat_id):
        WAITING_KEY.pop(update.message.chat_id)
        if not text.isdigit():
            await update.message.reply_text("Номер ключа должен быть числом")
            return
        key_number = int(text)
        df_full = full_stock(df)
        car = df_full[df_full["Номер ключа"] == key_number]
        if car.empty:
            await update.message.reply_text("Машина не найдена")
            return
        car = car.iloc[0]
        message = (
            f"🚗 Автомобиль найден\n\n"
            f"Номер ключа: {car['Номер ключа']}\n"
            f"Марка: {car['Марка']}\n"
            f"Модель: {car['Модель']}\n"
            f"VIN: {car['VIN']}\n"
            f"Год выпуска: {car['Год выпуска']}\n"
            f"Пробег: {car['Пробег']}\n"
            f"Цвет: {car['Цветкузова']}\n"
            f"Рег. номер: {car['Рег. номер']}\n"
            f"Парковка: {car['Место хранения']}\n"
            f"Дней в стоке: {car['Дней с даты поступления']}\n"
            f"Цена приема: {car['Цена приема']}\n"
            f"Цена продажи: {car['Цена продажи']}\n"
            f"Байер: {car['Байер']}\n"
            f"Тип сделки: {car['Тип сделки']}"
        )
        await update.message.reply_text(message)
        return

    # ===== Обработка сообщений из группы: регистрация ключей =====
    if chat_type in ["group", "supergroup"] and text:
        key_numbers = extract_key_numbers(text)
        if not key_numbers:
            return
        df_full = full_stock(df)
        keys_data = load_keys()
        responses = []
        for key_number in key_numbers:
            car_match = df_full[df_full["Номер ключа"] == key_number]
            if car_match.empty:
                responses.append(f"❌ Ключ {key_number} — не найден в стоке")
                logger.warning(f"Ключ {key_number} не найден в стоке")
                continue
            row = car_match.iloc[0]
            vin = str(row["VIN"]).strip() if "VIN" in row.index else ""
            marka = str(row.get("Марка", "")).strip()
            model = str(row.get("Модель", "")).strip()
            car_label = f"{marka} {model}".strip()
            existing = keys_data.get(str(key_number))
            if existing and existing.get("VIN") == vin:
                responses.append(f"ℹ️ Ключ {key_number} ({car_label}, VIN: {vin}) уже зарегистрирован")
            else:
                # Новая машина или переприсвоенный ключ — обновляем запись
                keys_data[str(key_number)] = {"VIN": vin, "status": "sent"}
                responses.append(f"✅ Ключ {key_number} — {car_label}\nVIN: {vin}")
                logger.info(f"Ключ {key_number} VIN={vin} {car_label} записан как переданный")
        save_keys(keys_data)
        if responses:
            await update.message.reply_text("\n\n".join(responses))
        return


# ===== Запуск бота =====
if __name__ == "__main__":
    if not TELEGRAM_TOKEN:
        logger.error("TELEGRAM_TOKEN не задан.")
        exit(1)

    app = ApplicationBuilder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler((filters.TEXT | filters.CAPTION) & ~filters.COMMAND, handle_text))
    logger.info("Бот запущен")
    app.run_polling()
